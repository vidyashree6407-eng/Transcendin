from openpyxl import load_workbook
wb = load_workbook('Courses & Categories - Final.xlsx', read_only=True)
print('sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    nonempty = [r for r in rows if any(cell is not None and str(cell).strip()!='' for cell in r)]
    print(name, 'rows_total=', len(rows), 'nonempty=', len(nonempty))
    for r in nonempty[:8]:
        print('  ', r)
