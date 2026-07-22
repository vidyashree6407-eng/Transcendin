from openpyxl import load_workbook
import csv

src = 'Courses & Categories - Final.xlsx'
out = 'c_c.csv'

wb = load_workbook(src, read_only=True)
ws = wb.active

with open(out, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    for row in ws.iter_rows(values_only=True):
        # convert None to empty string and ensure strings
        writer.writerow([str(cell).strip() if cell is not None else '' for cell in row])

print('Wrote', out)
